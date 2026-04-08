"""
NYC UAP / 485-x development strategy backend.
Uses Pinecone RAG + GPT multi-agent orchestration to evaluate zoning, site context, and developer-oriented scenarios.
"""

import os
import io
import re
import json
import logging
from contextlib import asynccontextmanager

from dotenv import load_dotenv
from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import anthropic
from openai import OpenAI
from pinecone import Pinecone, ServerlessSpec
from pydantic import BaseModel

from engine.engine import context_engine
from engine.helpers import helper_sanitize_input, helper_moderate_content, get_embedding, get_embeddings_batch
from property_models import (
    BlockLotsResponse,
    PropertyContext,
    PropertyContextRequest,
    PropertySearchResponse,
    ValidatedLotInfo,
)
from property_service import property_service
from property_store import delete_property_context, fetch_property_context, upsert_property_context
from tc201_models import TC201Data, ResidentialOccupancy, NonresidentialFloor, MiscExpenseItem
from underwriting_calculator import (
    build_underwriting_calculation_context,
    calculate_underwriting_formula_values,
    enable_workbook_recalculation,
)
from underwriting_template import build_underwriting_cell_payload

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "..", ".env"))

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# ── Configuration ───────────────────────────────────────────────────────

GENERATION_MODEL = os.getenv("GENERATION_MODEL", "claude-opus-4-6")
EMBEDDING_MODEL = os.getenv("EMBEDDING_MODEL", "text-embedding-3-large")
PINECONE_INDEX = os.getenv("PINECONE_INDEX", "und1")
NAMESPACE_CONTEXT = os.getenv("NAMESPACE_CONTEXT", "ContextLibrary")
NAMESPACE_KNOWLEDGE = os.getenv("NAMESPACE_KNOWLEDGE", "KnowledgeStore")
NAMESPACE_PROPERTY = os.getenv("NAMESPACE_PROPERTY", "PropertyContextStore")
DEFAULT_CORS_ALLOW_ORIGINS = ("http://localhost:3000", "http://localhost:3001", "http://127.0.0.1:35095", "http://localhost:35095")


def _get_csv_env(name: str, default: tuple[str, ...]) -> list[str]:
    raw_value = os.getenv(name, "")
    if not raw_value.strip():
        return list(default)
    values = [value.strip() for value in raw_value.split(",") if value.strip()]
    return values or list(default)


import re as _re

_DOMAIN_KEYWORDS = _re.compile(
    r"uap|485[\-\s]?x|zoning|far\b|floor area ratio|underwriting|rent roll|"
    r"t[\-\s]?12|noi\b|cap rate|debt service|dscr|affordable housing|"
    r"ami\b|hpd|hdc|nyc housing|tax abatement|tax exemption|"
    r"operating (expenses?|statement)|pro ?forma|rent stabiliz|"
    r"offering memorandum|appraisal|bbl\b|borough|lot area|building area|"
    r"development (strategy|scenario|site)|buildable|residential far|"
    r"commercial far|community facility",
    _re.IGNORECASE,
)

_SITE_KEYWORDS = _re.compile(
    r"\b(this|the|our|my)\s+(site|property|building|parcel|lot|project)\b",
    _re.IGNORECASE,
)


def _is_domain_query(query: str, context) -> bool:
    """Return True if *query* is within the NYC UAP / 485-x domain."""
    if _DOMAIN_KEYWORDS.search(query):
        return True
    if _SITE_KEYWORDS.search(query) and context is not None:
        return True
    return False


CORS_ALLOW_ORIGINS = _get_csv_env("CORS_ALLOW_ORIGINS", DEFAULT_CORS_ALLOW_ORIGINS)

# ── Global clients (initialized on startup) ────────────────────────────

openai_client: OpenAI | None = None
anthropic_client: anthropic.Anthropic | None = None
pinecone_client: Pinecone | None = None
active_index_name: str = PINECONE_INDEX  # mutable — switched via API
_template_store: dict = {}  # stores uploaded underwriting template bytes
_tc201_store: dict = {}  # stores parsed TC201 data

# Per-agent tunable settings
agent_settings: dict = {
    "librarian": {
        "top_k": 3,
        "description": "Retrieves semantic blueprints from the ContextLibrary to guide the Writer.",
    },
    "researcher": {
        "top_k": 20,
        "temperature": 0.1,
        "description": "Queries the KnowledgeStore and synthesizes facts with citations.",
    },
    "writer": {
        "temperature": 0.1,
        "description": "Combines research with the blueprint to generate the final response.",
    },
    "summarizer": {
        "temperature": 0.1,
        "max_length": 2000,
        "description": "Condenses large outputs to stay within token limits.",
    },
}


@asynccontextmanager
async def lifespan(_app: FastAPI):
    global openai_client, anthropic_client, pinecone_client, active_index_name

    openai_key = os.getenv("OPENAI_API_KEY", "")
    anthropic_key = os.getenv("ANTHROPIC_API_KEY", "")
    pinecone_key = os.getenv("PINECONE_API_KEY", "")

    if not openai_key:
        logging.error("OPENAI_API_KEY not set!")
    if not anthropic_key:
        logging.error("ANTHROPIC_API_KEY not set!")
    if not pinecone_key:
        logging.error("PINECONE_API_KEY not set!")

    openai_client = OpenAI(api_key=openai_key)
    anthropic_client = anthropic.Anthropic(api_key=anthropic_key)
    pinecone_client = Pinecone(api_key=pinecone_key)

    # Verify Pinecone index exists; auto-switch if the configured one is gone
    try:
        idx = pinecone_client.Index(active_index_name)
        stats = idx.describe_index_stats()
        total_vectors = stats.get("total_vector_count", 0)
        logging.info(f"✅ Pinecone connected: index={active_index_name}, vectors={total_vectors}")
    except Exception as e:
        logging.warning(f"Configured index '{active_index_name}' not reachable: {e}")
        # Try to fall back to the first available index
        try:
            available = pinecone_client.list_indexes()
            if available:
                first = available[0].name
                active_index_name = first
                idx = pinecone_client.Index(active_index_name)
                stats = idx.describe_index_stats()
                total_vectors = stats.get("total_vector_count", 0)
                logging.info(f"✅ Auto-switched to index '{active_index_name}', vectors={total_vectors}")
            else:
                logging.warning("No Pinecone indexes available")
        except Exception as e2:
            logging.error(f"Failed to auto-switch index: {e2}")

    logging.info(f"✅ MAS Backend ready — model={GENERATION_MODEL}, embedding={EMBEDDING_MODEL}")
    yield


app = FastAPI(title="UAP 485-x NYC Development Expert API", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ALLOW_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Models ──────────────────────────────────────────────────────────────

class ChatMessage(BaseModel):
    role: str
    content: str


class ChatRequest(BaseModel):
    messages: list[ChatMessage]
    use_rag: bool = True


class ChatResponse(BaseModel):
    reply: str
    sources: list[dict]


class UnderwritingUpdatesRequest(BaseModel):
    updates: dict[str, dict[str, str | int | float]] = {}


def _get_index():
    return pinecone_client.Index(active_index_name)


def _get_active_property_context() -> PropertyContext | None:
    try:
        return fetch_property_context(_get_index(), NAMESPACE_PROPERTY)
    except Exception as exc:
        logging.warning(f"Failed to load active property context: {exc}")
        return None


# Borough name → code mapping for BBL construction
_BOROUGH_CODES = {
    "manhattan": "1", "bronx": "2", "brooklyn": "3", "queens": "4", "statenisland": "5",
    "staten island": "5", "si": "5", "bk": "3", "bx": "2", "mn": "1", "qn": "4",
}


def _parse_bbl_from_filename(filename: str) -> str | None:
    """Extract a 10-digit BBL from filename pattern (Borough_Block_Lot).
    Examples:
        'ESR EQUITY LLC_Profit and Loss (Queens_2090_3).xlsx' → '4020900003'
        '2025 Ocean Avenue Condominium - 2025 (Brooklyn_6767_1301-1316).pdf' → '3067670001301'
    Returns the primary BBL (first lot if range) or None.
    """
    m = re.search(r'\(([A-Za-z ]+?)_(\d+)_(\d+)(?:-\d+)?\)', filename)
    if not m:
        return None
    borough_name, block_str, lot_str = m.group(1), m.group(2), m.group(3)
    borough_code = _BOROUGH_CODES.get(borough_name.strip().lower())
    if not borough_code:
        return None
    bbl = f"{borough_code}{block_str.zfill(5)}{lot_str.zfill(4)}"
    if len(bbl) != 10:
        return None
    return bbl


async def _auto_set_property_context(bbl: str) -> PropertyContext | None:
    """Auto-set property context from a BBL. Returns the context or None on failure."""
    try:
        existing = _get_active_property_context()
        if existing and existing.primary_bbl == bbl:
            logging.info(f"Property context already set for BBL {bbl}, skipping auto-set")
            return existing
        context = await property_service.build_property_context(bbl, [])
        embedding = get_embedding(context.property_brief, client=openai_client, embedding_model=EMBEDDING_MODEL)
        upsert_property_context(_get_index(), NAMESPACE_PROPERTY, embedding, context)
        logging.info(f"Auto-set property context for BBL {bbl}: {context.address}")
        return context
    except Exception as exc:
        logging.warning(f"Auto-set property context failed for BBL {bbl}: {exc}")
        return None


# ── Agent Settings ──────────────────────────────────────────────────────

@app.get("/api/settings")
async def get_settings():
    """Return current agent settings."""
    return {"settings": agent_settings}


@app.put("/api/settings")
async def update_settings(req: dict):
    """Update agent settings. Accepts partial updates per agent."""
    ALLOWED_KEYS = {
        "librarian": {"top_k": (int, 1, 20)},
        "researcher": {"top_k": (int, 1, 100), "temperature": (float, 0.0, 2.0)},
        "writer": {"temperature": (float, 0.0, 2.0)},
        "summarizer": {"temperature": (float, 0.0, 2.0), "max_length": (int, 100, 10000)},
    }
    settings_input = req.get("settings", {})
    for agent_name, params in settings_input.items():
        if agent_name not in ALLOWED_KEYS:
            continue
        for key, value in params.items():
            if key not in ALLOWED_KEYS[agent_name]:
                continue
            expected_type, min_val, max_val = ALLOWED_KEYS[agent_name][key]
            try:
                casted = expected_type(value)
                casted = max(min_val, min(max_val, casted))
                agent_settings[agent_name][key] = casted
            except (ValueError, TypeError):
                continue
    logging.info(f"Agent settings updated: {agent_settings}")
    return {"settings": agent_settings}


# ── Pinecone Index Management ───────────────────────────────────────────

class CreateIndexRequest(BaseModel):
    name: str
    dimension: int = 3072
    metric: str = "cosine"
    cloud: str = "aws"
    region: str = "us-east-1"


class SwitchIndexRequest(BaseModel):
    name: str


@app.get("/api/indexes")
async def list_indexes():
    """List all Pinecone indexes in the account."""
    try:
        indexes = pinecone_client.list_indexes()
        result = []
        for idx_model in indexes:
            status = idx_model.status
            result.append({
                "name": idx_model.name,
                "dimension": idx_model.dimension,
                "metric": idx_model.metric,
                "host": idx_model.host,
                "ready": status["ready"] if status else False,
                "state": status["state"] if status else "Unknown",
            })
        return {"indexes": result, "active": active_index_name}
    except Exception as e:
        logging.error(f"Failed to list indexes: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/indexes/active")
async def get_active_index():
    """Return the currently active Pinecone index with stats."""
    try:
        idx = pinecone_client.Index(active_index_name)
        stats = idx.describe_index_stats()
        ns = stats.get("namespaces", {})
        return {
            "name": active_index_name,
            "total_vectors": stats.get("total_vector_count", 0),
            "dimension": stats.get("dimension", 0),
            "namespaces": {
                k: {"vector_count": v.get("vector_count", 0)}
                for k, v in ns.items()
            },
        }
    except Exception as e:
        logging.error(f"Failed to get active index stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/indexes")
async def create_index(req: CreateIndexRequest):
    """Create a new Pinecone serverless index."""
    try:
        pinecone_client.create_index(
            name=req.name,
            dimension=req.dimension,
            metric=req.metric,
            spec=ServerlessSpec(cloud=req.cloud, region=req.region),
        )
        logging.info(f"Created Pinecone index: {req.name}")
        return {"created": req.name}
    except Exception as e:
        logging.error(f"Failed to create index: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/indexes/{index_name}")
async def delete_index(index_name: str):
    """Delete a Pinecone index."""
    if index_name == active_index_name:
        raise HTTPException(status_code=400, detail="Cannot delete the currently active index. Switch to another first.")
    try:
        pinecone_client.delete_index(name=index_name)
        logging.info(f"Deleted Pinecone index: {index_name}")
        return {"deleted": index_name}
    except Exception as e:
        logging.error(f"Failed to delete index: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/indexes/switch")
async def switch_index(req: SwitchIndexRequest):
    """Switch the active Pinecone index used for chat and uploads."""
    global active_index_name
    # Verify the index exists and is ready
    try:
        info = pinecone_client.describe_index(req.name)
        status = info.status
        if not status["ready"]:
            raise HTTPException(status_code=400, detail=f"Index '{req.name}' is not ready")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"Index '{req.name}' not found: {e}")

    active_index_name = req.name
    logging.info(f"Switched active index to: {active_index_name}")

    # Auto-detect property from existing documents if no context is set
    auto_property = None
    existing_ctx = _get_active_property_context()
    if not existing_ctx:
        try:
            idx = pinecone_client.Index(active_index_name)
            for page in idx.list(namespace=NAMESPACE_KNOWLEDGE):
                for item in page:
                    vid = item if isinstance(item, str) else getattr(item, "id", str(item))
                    bbl = _parse_bbl_from_filename(vid)
                    if bbl:
                        auto_property = await _auto_set_property_context(bbl)
                        break
                if auto_property:
                    break
        except Exception as exc:
            logging.warning(f"Auto-detect property on index switch failed: {exc}")

    result = {"active": active_index_name}
    if auto_property:
        result["auto_property"] = {
            "bbl": auto_property.primary_bbl,
            "address": auto_property.address,
        }
    return result


# ── Live Property Context ───────────────────────────────────────────────

@app.get("/api/property/search-address", response_model=PropertySearchResponse)
async def search_property_address(q: str = Query("", description="NYC address or 10-digit BBL")):
    try:
        results = await property_service.search_address(q)
        return PropertySearchResponse(results=results, query=q)
    except Exception as exc:
        logging.error(f"Property address search failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Address search failed: {exc}")


@app.get("/api/property/validate-lot", response_model=ValidatedLotInfo)
async def validate_property_lot(bbl: str = Query(..., description="10-digit BBL")):
    try:
        result = await property_service.validate_lot(bbl)
        return ValidatedLotInfo(**result)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        logging.error(f"Property lot validation failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Lot validation failed: {exc}")


@app.get("/api/property/block-lots", response_model=BlockLotsResponse)
async def get_property_block_lots(
    borough: int = Query(..., ge=1, le=5, description="Borough code (1-5)"),
    block: int = Query(..., gt=0, description="Tax block"),
):
    try:
        result = await property_service.get_block_lots(borough, block)
        return BlockLotsResponse(**result)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        logging.error(f"Property block lookup failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Block lot lookup failed: {exc}")


@app.put("/api/property/context", response_model=PropertyContext)
async def set_property_context(req: PropertyContextRequest):
    try:
        context = await property_service.build_property_context(req.primary_bbl, req.adjacent_bbls)
        embedding = get_embedding(context.property_brief, client=openai_client, embedding_model=EMBEDDING_MODEL)
        upsert_property_context(_get_index(), NAMESPACE_PROPERTY, embedding, context)
        logging.info(f"Stored active property context for index '{active_index_name}': {context.primary_bbl}")
        return context
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        logging.error(f"Set property context failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Property context update failed: {exc}")


@app.get("/api/property/context", response_model=PropertyContext | None)
async def get_property_context():
    try:
        return _get_active_property_context()
    except Exception as exc:
        logging.error(f"Get property context failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Property context lookup failed: {exc}")


@app.delete("/api/property/context")
async def clear_property_context():
    try:
        delete_property_context(_get_index(), NAMESPACE_PROPERTY)
        logging.info(f"Cleared active property context for index '{active_index_name}'")
        return {"cleared": True}
    except Exception as exc:
        logging.error(f"Clear property context failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Property context clear failed: {exc}")


# ── Blueprint Management (ContextLibrary) ───────────────────────────────

class CreateBlueprintRequest(BaseModel):
    subject: str
    instructions: str


@app.get("/api/blueprints")
async def list_blueprints():
    """List all blueprints in the ContextLibrary namespace."""
    try:
        idx = pinecone_client.Index(active_index_name)
        stats = idx.describe_index_stats()
        dim = stats.get("dimension", 3072) or 3072
        # Fetch all vectors in ContextLibrary using a zero-vector query
        # (Pinecone doesn't have a "list" — we query with a dummy and high top_k)
        dummy_vec = [0.0] * int(dim)
        results = idx.query(
            vector=dummy_vec,
            top_k=100,
            namespace=NAMESPACE_CONTEXT,
            include_metadata=True,
        )
        blueprints = []
        for m in results.get("matches", []):
            meta = m.get("metadata", {})
            blueprints.append({
                "id": m["id"],
                "subject": meta.get("subject", "Unknown"),
                "instructions": meta.get("text", ""),
            })
        return {"blueprints": blueprints}
    except Exception as e:
        logging.error(f"Failed to list blueprints: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/blueprints")
async def create_blueprint(req: CreateBlueprintRequest):
    """Create a new blueprint in the ContextLibrary namespace."""
    if not req.subject.strip() or not req.instructions.strip():
        raise HTTPException(status_code=400, detail="Subject and instructions are required")

    try:
        idx = pinecone_client.Index(active_index_name)
        # Embed the subject so the Librarian can match it semantically
        embedding = get_embedding(req.subject.strip(), client=openai_client, embedding_model=EMBEDDING_MODEL)
        vec_id = f"blueprint__{req.subject.strip().lower().replace(' ', '_')}"
        idx.upsert(
            vectors=[{
                "id": vec_id,
                "values": embedding,
                "metadata": {
                    "text": req.instructions.strip(),
                    "subject": req.subject.strip(),
                    "source": "blueprint",
                },
            }],
            namespace=NAMESPACE_CONTEXT,
        )
        logging.info(f"Created blueprint: {req.subject}")
        return {"id": vec_id, "subject": req.subject.strip()}
    except Exception as e:
        logging.error(f"Failed to create blueprint: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class GenerateBlueprintRequest(BaseModel):
    subject: str


@app.post("/api/blueprints/generate")
async def generate_blueprint(req: GenerateBlueprintRequest):
    """Use AI to generate blueprint instructions for a subject, then store it."""
    subject = req.subject.strip()
    if not subject:
        raise HTTPException(status_code=400, detail="Subject is required")

    try:
        # Generate instructions via GPT
        system_prompt = (
            "You are writing blueprint instructions for an AI Writer that is exclusively focused on "
            "NYC UAP / 485-x building development strategy. Given a subject domain, produce instructions "
            "for how the Writer should format, structure, and tone responses inside that NYC development context.\n\n"
            "Include: developer-first tone, response structure, terminology guidance, profitability framing, "
            "assumption handling, and how to cite source-grounded constraints.\n\n"
            "Be specific and practical. Output ONLY the instructions, no preamble."
        )
        response = anthropic_client.messages.create(
            model=GENERATION_MODEL,
            max_tokens=4096,
            temperature=0.4,
            system=system_prompt,
            messages=[
                {"role": "user", "content": f"Generate blueprint instructions for the subject: {subject}"},
            ],
        )
        instructions = response.content[0].text.strip()

        # Upsert into Pinecone ContextLibrary
        idx = pinecone_client.Index(active_index_name)
        embedding = get_embedding(subject, client=openai_client, embedding_model=EMBEDDING_MODEL)
        vec_id = f"blueprint__{subject.lower().replace(' ', '_')}"
        idx.upsert(
            vectors=[{
                "id": vec_id,
                "values": embedding,
                "metadata": {
                    "text": instructions,
                    "subject": subject,
                    "source": "blueprint",
                },
            }],
            namespace=NAMESPACE_CONTEXT,
        )
        logging.info(f"AI-generated blueprint for '{subject}'")
        return {"id": vec_id, "subject": subject, "instructions": instructions}
    except Exception as e:
        logging.error(f"Failed to generate blueprint: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/blueprints/{blueprint_id:path}")
async def delete_blueprint(blueprint_id: str):
    """Delete a blueprint from the ContextLibrary namespace."""
    try:
        idx = pinecone_client.Index(active_index_name)
        idx.delete(ids=[blueprint_id], namespace=NAMESPACE_CONTEXT)
        logging.info(f"Deleted blueprint: {blueprint_id}")
        return {"deleted": blueprint_id}
    except Exception as e:
        logging.error(f"Failed to delete blueprint: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Routes ──────────────────────────────────────────────────────────────

@app.get("/api/health")
async def health():
    try:
        idx = pinecone_client.Index(active_index_name)
        stats = idx.describe_index_stats()
        total = stats.get("total_vector_count", 0)
    except Exception:
        total = 0
    return {"status": "ok", "documents": total, "active_index": active_index_name}


@app.post("/api/chat", response_model=ChatResponse)
async def chat(req: ChatRequest):
    # Extract the last user message as the goal
    last_user_msg = ""
    for m in reversed(req.messages):
        if m.role == "user":
            last_user_msg = m.content
            break

    if not last_user_msg:
        raise HTTPException(status_code=400, detail="No user message found")

    # Sanitize and moderate input
    try:
        sanitized = helper_sanitize_input(last_user_msg)
        helper_moderate_content(sanitized, openai_client)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    property_context = _get_active_property_context()

    # Build conversation history for multi-turn context
    conversation_history = [
        {"role": m.role, "content": m.content}
        for m in req.messages[:-1]  # Exclude last message (it's the goal)
    ] if len(req.messages) > 1 else None

    # Run the Multi-Agent System pipeline
    try:
        result, trace = context_engine(
            goal=sanitized,
            client=anthropic_client,
            pc=pinecone_client,
            index_name=active_index_name,
            generation_model=GENERATION_MODEL,
            embedding_model=EMBEDDING_MODEL,
            namespace_context=NAMESPACE_CONTEXT,
            namespace_knowledge=NAMESPACE_KNOWLEDGE,
            agent_settings=agent_settings,
            property_context=property_context.model_dump() if property_context else None,
            conversation_history=conversation_history,
            embedding_client=openai_client,
        )
    except Exception as e:
        logging.error(f"Context engine error: {e}")
        raise HTTPException(status_code=500, detail="Engine processing failed")

    if result is None:
        return ChatResponse(
            reply=f"I wasn't able to find a complete answer. Trace: {trace.status}",
            sources=[],
        )

    # Extract final text and sources from the MAS output
    if isinstance(result, str):
        reply_text = result
        sources = []
    elif isinstance(result, dict):
        reply_text = result.get("answer_with_sources", result.get("summary", str(result)))
        sources = result.get("sources", [])
    else:
        reply_text = str(result)
        sources = []

    # Format sources for the frontend contract
    formatted_sources = []
    for s in sources[:10]:
        if isinstance(s, dict):
            source_name = s.get("source", "Pinecone")
            formatted_sources.append({
                "filename": source_name,
                "distance": round(1 - s.get("score", 0), 4),
                "source_type": "property" if str(source_name).startswith("Active Property Context") else "document",
            })

    return ChatResponse(reply=reply_text, sources=formatted_sources)


@app.post("/api/chat/stream")
async def chat_stream(req: ChatRequest):
    """Streaming chat endpoint — returns Server-Sent Events (SSE)."""
    from engine.agents import researcher_stream, agent_context_librarian
    from engine.helpers import create_mcp_message as _mcp_msg

    last_user_msg = ""
    for m in reversed(req.messages):
        if m.role == "user":
            last_user_msg = m.content
            break

    if not last_user_msg:
        raise HTTPException(status_code=400, detail="No user message found")

    try:
        sanitized = helper_sanitize_input(last_user_msg)
        helper_moderate_content(sanitized, openai_client)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    property_context = _get_active_property_context()
    conversation_history = [
        {"role": m.role, "content": m.content}
        for m in req.messages[:-1]
    ] if len(req.messages) > 1 else None

    def event_generator():
        try:
            # Step 1: Librarian (fast vector search for blueprint)
            idx = _get_index()
            lib_msg = _mcp_msg("Engine", {"intent_query": sanitized})
            lib_result = agent_context_librarian(
                lib_msg, client=openai_client, index=idx,
                embedding_model=EMBEDDING_MODEL, namespace_context=NAMESPACE_CONTEXT,
                agent_settings=agent_settings,
            )
            blueprint_text = lib_result["content"].get("blueprint_json", "")

            # Step 2: Stream Researcher output
            for event_type, data in researcher_stream(
                goal=sanitized,
                client=anthropic_client,
                index=idx,
                generation_model=GENERATION_MODEL,
                embedding_model=EMBEDDING_MODEL,
                namespace_knowledge=NAMESPACE_KNOWLEDGE,
                agent_settings=agent_settings,
                property_context=property_context.model_dump() if property_context else None,
                conversation_history=conversation_history,
                blueprint_text=blueprint_text,
                embedding_client=openai_client,
            ):
                if event_type == "sources":
                    formatted = []
                    for s in data:
                        if isinstance(s, dict):
                            source_name = s.get("source", "Pinecone")
                            formatted.append({
                                "filename": source_name,
                                "distance": round(1 - s.get("score", 0), 4),
                                "source_type": "property" if str(source_name).startswith("Active Property Context") else "document",
                            })
                    yield f"data: {json.dumps({'type': 'sources', 'sources': formatted})}\n\n"
                elif event_type == "chunk":
                    yield f"data: {json.dumps({'type': 'chunk', 'text': data})}\n\n"

            yield f"data: {json.dumps({'type': 'done'})}\n\n"

        except Exception as e:
            logging.error(f"Stream error: {e}")
            yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/upload")
async def upload_document(file: UploadFile = File(...)):
    """Upload any document and upsert its chunks into Pinecone KnowledgeStore."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    content = await file.read()
    ext = os.path.splitext(file.filename)[1].lower()

    # Extract text based on file type
    text = _extract_text(content, ext, file.filename)

    if len(text.strip()) == 0:
        raise HTTPException(status_code=400, detail="File is empty or could not be parsed")

    # Chunk the text
    chunks = _chunk_text(text, chunk_size=800, overlap=200)
    if not chunks:
        raise HTTPException(status_code=400, detail="No chunks generated")

    # Upsert into Pinecone
    idx = pinecone_client.Index(active_index_name)
    embeddings = get_embeddings_batch(chunks, client=openai_client, embedding_model=EMBEDDING_MODEL)
    vectors = []
    for i, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
        vec_id = f"{file.filename}__chunk_{i}"
        vectors.append({
            "id": vec_id,
            "values": embedding,
            "metadata": {"text": chunk, "source": file.filename, "chunk_index": i},
        })

    # Upsert in batches of 100
    for batch_start in range(0, len(vectors), 100):
        batch = vectors[batch_start:batch_start + 100]
        idx.upsert(vectors=batch, namespace=NAMESPACE_KNOWLEDGE)

    logging.info(f"Uploaded {len(chunks)} chunks for '{file.filename}' to Pinecone.")

    # Auto-set property context from filename BBL pattern
    auto_property = None
    parsed_bbl = _parse_bbl_from_filename(file.filename)
    if parsed_bbl:
        logging.info(f"Parsed BBL {parsed_bbl} from filename '{file.filename}'")
        auto_property = await _auto_set_property_context(parsed_bbl)

    result = {"filename": file.filename, "chunks": len(chunks)}
    if auto_property:
        result["auto_property"] = {
            "bbl": auto_property.primary_bbl,
            "address": auto_property.address,
        }
    return result


@app.get("/api/documents")
async def list_documents():
    """List uploaded documents (by filename) from the active index's KnowledgeStore."""
    try:
        idx = pinecone_client.Index(active_index_name)
        filenames: dict[str, int] = {}  # filename → chunk count

        for page in idx.list(namespace=NAMESPACE_KNOWLEDGE):
            for item in page:
                vid = item if isinstance(item, str) else getattr(item, "id", str(item))
                if "__chunk_" in vid:
                    fname = vid.rsplit("__chunk_", 1)[0]
                    filenames[fname] = filenames.get(fname, 0) + 1

        docs = [{"filename": f, "chunks": c} for f, c in sorted(filenames.items())]
        total = sum(filenames.values())
        return {"documents": docs, "total_chunks": total}
    except Exception as e:
        logging.error(f"List documents failed: {e}")
        return {"documents": [], "total_chunks": 0}


@app.delete("/api/documents/{filename:path}")
async def delete_document(filename: str):
    """Delete all vectors for a given source filename from Pinecone."""
    try:
        idx = pinecone_client.Index(active_index_name)
        # Delete by metadata filter
        idx.delete(
            filter={"source": {"$eq": filename}},
            namespace=NAMESPACE_KNOWLEDGE,
        )
        logging.info(f"Deleted vectors for '{filename}' from Pinecone.")
        return {"deleted_chunks": 1}  # Pinecone doesn't return count
    except Exception as e:
        logging.error(f"Delete failed: {e}")
        raise HTTPException(status_code=500, detail=f"Delete failed: {e}")


def _extract_text(content: bytes, ext: str, filename: str) -> str:
    """Extract plain text from any supported file type."""
    # PDF — prefer pdfplumber for table-aware extraction, fall back to pypdf
    if ext == ".pdf":
        try:
            import pdfplumber
            pages_text: list[str] = []
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for page in pdf.pages:
                    # Extract tables first so we can represent them structurally
                    tables = page.extract_tables() or []
                    table_texts: list[str] = []
                    for table in tables:
                        if not table:
                            continue
                        header = table[0]
                        for row in table[1:]:
                            if not row:
                                continue
                            row_parts: list[str] = []
                            for idx, cell in enumerate(row):
                                col_name = header[idx] if header and idx < len(header) and header[idx] else f"Col{idx+1}"
                                cell_val = (cell or "").strip()
                                if cell_val:
                                    row_parts.append(f"{col_name}: {cell_val}")
                            if row_parts:
                                table_texts.append(" | ".join(row_parts))

                    # Get non-table text
                    page_text = page.extract_text() or ""
                    if table_texts:
                        page_text += "\n\n[Table data]\n" + "\n".join(table_texts)
                    pages_text.append(page_text)
            return "\n\n".join(pages_text)
        except ImportError:
            pass  # fall through to pypdf
        except Exception as e:
            logging.warning(f"pdfplumber failed for {filename}, falling back to pypdf: {e}")

        # Fallback: pypdf
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(content))
            pages = [page.extract_text() or "" for page in reader.pages]
            return "\n\n".join(pages)
        except Exception as e:
            logging.error(f"PDF parse failed for {filename}: {e}")
            raise HTTPException(status_code=400, detail=f"Failed to parse PDF: {e}")

    # DOCX
    if ext == ".docx":
        try:
            from docx import Document
            doc = Document(io.BytesIO(content))
            return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except Exception as e:
            logging.error(f"DOCX parse failed for {filename}: {e}")
            raise HTTPException(status_code=400, detail=f"Failed to parse DOCX: {e}")

    # XLSX / XLS — structured extraction preserving column headers per row
    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            lines: list[str] = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"=== Sheet: {sheet} ===")
                header_cells: list[str] = []
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if row_idx == 0:
                        header_cells = [str(c).strip() if c is not None else f"Col{i+1}" for i, c in enumerate(row)]
                        lines.append("Columns: " + " | ".join(header_cells))
                        continue
                    parts: list[str] = []
                    for col_idx, cell in enumerate(row):
                        if cell is None:
                            continue
                        col_name = header_cells[col_idx] if col_idx < len(header_cells) else f"Col{col_idx+1}"
                        parts.append(f"{col_name}: {cell}")
                    if parts:
                        lines.append(" | ".join(parts))
            return "\n".join(lines)
        except Exception as e:
            logging.error(f"Excel parse failed for {filename}: {e}")
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {e}")

    # Everything else — treat as UTF-8 text
    return content.decode("utf-8", errors="replace")


def _chunk_text(text: str, chunk_size: int = 800, overlap: int = 200) -> list[str]:
    """Split text into overlapping chunks, preferring paragraph boundaries."""
    import re
    # Split into paragraphs first (double newlines, section headers, page breaks)
    paragraphs = re.split(r'\n{2,}|(?=^={3,}|^---)', text, flags=re.MULTILINE)
    paragraphs = [p.strip() for p in paragraphs if p.strip()]

    chunks: list[str] = []
    current_chunk = ""

    for para in paragraphs:
        # If a single paragraph exceeds chunk_size, sub-split it the old way
        if len(para) > chunk_size:
            # Flush what we have
            if current_chunk.strip():
                chunks.append(current_chunk.strip())
                current_chunk = ""
            start = 0
            while start < len(para):
                end = start + chunk_size
                # Try to break at a sentence or line boundary
                if end < len(para):
                    boundary = max(
                        para.rfind(". ", start + chunk_size // 2, end),
                        para.rfind("\n", start + chunk_size // 2, end),
                    )
                    if boundary > start:
                        end = boundary + 1
                chunks.append(para[start:end].strip())
                start = end - overlap if end - overlap > start else end
            continue

        # Would adding this paragraph exceed chunk_size?
        candidate = (current_chunk + "\n\n" + para).strip() if current_chunk else para
        if len(candidate) <= chunk_size:
            current_chunk = candidate
        else:
            if current_chunk.strip():
                chunks.append(current_chunk.strip())
            # Start new chunk with overlap from the end of previous chunk
            if overlap > 0 and chunks:
                tail = chunks[-1][-overlap:]
                current_chunk = tail + "\n\n" + para
            else:
                current_chunk = para

    if current_chunk.strip():
        chunks.append(current_chunk.strip())

    return [c for c in chunks if c.strip()]


def _select_diversified_source_chunks(
    matches: list[dict],
    *,
    max_sources: int = 10,
    max_chunks_per_source: int = 8,
    max_total_chunks: int = 40,
) -> tuple[list[str], dict[str, str]]:
    """Balance retrieved chunks across documents so one file does not dominate extraction."""
    chunks_by_source: dict[str, list[tuple[int | None, str]]] = {}
    source_order: list[str] = []

    for match in matches:
        metadata = match.get("metadata", {}) or {}
        chunk_text = metadata.get("text", "")
        source_name = str(metadata.get("source", "Unknown source")).strip() or "Unknown source"
        if not chunk_text:
            continue

        if source_name not in chunks_by_source:
            chunks_by_source[source_name] = []
            source_order.append(source_name)

        if len(chunks_by_source[source_name]) >= max_chunks_per_source:
            continue

        raw_chunk_index = metadata.get("chunk_index")
        chunk_index = raw_chunk_index if isinstance(raw_chunk_index, int) else None
        chunks_by_source[source_name].append((chunk_index, chunk_text))

    selected_sources = source_order[:max_sources]
    source_chunks: list[str] = []
    source_name_lookup: dict[str, str] = {}

    round_index = 0
    while len(source_chunks) < max_total_chunks:
        added_any = False
        for source_name in selected_sources:
            chunks = chunks_by_source.get(source_name, [])
            if round_index >= len(chunks):
                continue

            chunk_index, chunk_text = chunks[round_index]
            header = f"[Source: {source_name}]"
            if chunk_index is not None:
                header += f"\n[Chunk {chunk_index}]"
            source_chunks.append(f"{header}\n{chunk_text}")
            source_name_lookup[source_name.lower()] = source_name
            added_any = True

            if len(source_chunks) >= max_total_chunks:
                break

        if not added_any:
            break
        round_index += 1

    return source_chunks, source_name_lookup


def _pick_relevant_glossary_sections(labels: list[str], sheet_name: str) -> str:
    """Select only the glossary sections relevant to this sheet's labels."""
    from underwriting_domain import UNDERWRITING_GLOSSARY

    sheet_lower = sheet_name.lower()
    label_blob = " ".join(labels).lower()
    combined = f"{sheet_lower} {label_blob}"

    section_keywords: dict[str, list[str]] = {
        "Property Information": ["address", "bbl", "block", "lot", "zoning", "far", "buildable", "height", "borough", "overlay"],
        "Unit Mix": ["unit", "studio", "1br", "2br", "3br", "bedroom", "ami", "affordable", "market rate", "dwelling", "duf"],
        "Revenue": ["rent", "gpr", "egi", "vacancy", "income", "laundry", "parking", "storage", "commercial", "collection"],
        "Operating Expenses": ["expense", "opex", "tax", "insurance", "payroll", "management", "repair", "utility", "water", "electric", "gas", "fuel", "elevator", "r&m", "reserve"],
        "NOI & Valuation": ["noi", "cap rate", "valuation", "appraised", "price per", "grm"],
        "Acquisition & Financing": ["purchase", "acquisition", "closing", "transfer", "hard cost", "soft cost", "tdc", "equity", "ltv", "ltc"],
        "Debt Service": ["mortgage", "loan", "interest", "amortization", "debt service", "dscr", "construction loan", "permanent", "mezzanine"],
        "Returns": ["cash flow", "irr", "cash-on-cash", "coc", "equity multiple", "hold period", "exit cap", "reversion", "roi", "yield", "btcf", "atcf"],
        "Tax Programs": ["uap", "485", "421", "abatement", "icap", "j-51", "prevailing", "pilot"],
        "Development": ["gsf", "nsf", "rsf", "efficiency", "stories", "floor", "cellar", "parking", "construction type", "lease-up", "stabilization"],
        "Rent Regulation": ["dhcr", "rgb", "mci", "iai", "hstpa", "stabilized", "regulated", "decontrol"],
        "Sources & Uses": ["source", "uses", "land cost", "contingency", "developer fee", "lihtc", "hpd subsidy", "hdc bond", "gap funding", "interest reserve"],
        "Pro Forma Projections": ["year 1", "year 2", "pro forma", "rent growth", "expense growth", "npv", "discount rate", "terminal", "levered irr"],
        "Sensitivity Analysis": ["base case", "downside", "upside", "break-even", "stress", "sensitivity"],
        "Deal Structure": ["gp", "lp", "sponsor", "preferred return", "promote", "waterfall", "catch-up", "capital stack", "joint venture"],
        "485-x Program": ["485-x", "benefit period", "phase-out", "affordability lock", "prevailing wage", "regulatory agreement", "hpd marketing"],
        "UAP Scenarios": ["as-of-right", "full bonus", "avoid prevailing", "avoid 40%", "bonus floor", "optimized"],
        "NYC Compliance & Expenses": ["scrie", "drie", "hpd violation", "lead paint", "local law", "ll97", "ll11", "ll87", "landmark", "certificate of occupancy", "dob", "sro"],
        "Abbreviations": [],
    }

    selected_sections: list[str] = []
    for section, keywords in section_keywords.items():
        if section == "Abbreviations":
            selected_sections.append(section)
            continue
        if any(kw in combined for kw in keywords):
            selected_sections.append(section)

    for core in ["Property Information", "Unit Mix", "Revenue", "Operating Expenses", "NOI & Valuation"]:
        if core not in selected_sections:
            selected_sections.append(core)

    lines: list[str] = ["DOMAIN GLOSSARY (relevant sections):\n"]
    for section in selected_sections:
        entries = UNDERWRITING_GLOSSARY.get(section, [])
        if not entries:
            continue
        lines.append(f"## {section}")
        for label, meaning in entries:
            lines.append(f"  - {label}: {meaning}")
        lines.append("")
    return "\n".join(lines)


# ── Underwriting Template ───────────────────────────────────────────────

def _safe_cell_value(val):
    """Convert cell value to a JSON-serializable type."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, float):
        if val != val:  # NaN
            return None
        return round(val, 2)
    if isinstance(val, int):
        return val
    if isinstance(val, str):
        return val
    return str(val)


def _col_letter(col: int) -> str:
    """Convert 1-based column number to Excel column letter(s)."""
    result = ""
    while col > 0:
        col -= 1
        result = chr(65 + col % 26) + result
        col //= 26
    return result


@app.post("/api/underwriting/parse-template")
async def parse_underwriting_template(file: UploadFile = File(...)):
    """Upload an Excel underwriting template and return its parsed structure."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    content = await file.read()
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".xlsx", ".xls"):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx / .xls) are supported")

    import openpyxl

    wb_vals = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    wb_fmls = openpyxl.load_workbook(io.BytesIO(content), data_only=False)

    sheets = []
    formula_refs_by_sheet: dict[str, list[str]] = {}
    for name in wb_vals.sheetnames:
        ws_v = wb_vals[name]
        ws_f = wb_fmls[name]
        max_r = ws_v.max_row or 0
        max_c = ws_v.max_column or 0

        data = []
        formula_refs: list[str] = []
        for r in range(1, max_r + 1):
            row = []
            for c in range(1, max_c + 1):
                value_cell = ws_v.cell(r, c)
                formula_cell = ws_f.cell(r, c)
                fml = formula_cell.value
                is_formula = isinstance(fml, str) and fml.startswith("=")
                cell = build_underwriting_cell_payload(
                    value_cell.value,
                    row=r,
                    col=c,
                    is_formula=is_formula,
                    number_format=formula_cell.number_format,
                    epoch=wb_vals.epoch,
                )
                if cell is None:
                    row.append(None)
                else:
                    if is_formula:
                        formula_refs.append(f"{_col_letter(c)}{r}")
                    row.append(cell)
            data.append(row)

        sheets.append({"name": name, "data": data, "maxRow": max_r, "maxCol": max_c})
        formula_refs_by_sheet[name] = formula_refs

    calc_context = build_underwriting_calculation_context(
        file.filename,
        content,
        formula_refs_by_sheet,
    )
    _template_store["current"] = {
        "filename": file.filename,
        "bytes": content,
        "formula_refs_by_sheet": formula_refs_by_sheet,
        "calc_context": calc_context,
    }
    logging.info(f"Parsed underwriting template: {file.filename} ({len(sheets)} sheets)")
    return {"filename": file.filename, "sheets": sheets}


@app.post("/api/underwriting/extract")
async def extract_underwriting_values():
    """Use RAG to extract values from uploaded source documents for each template sheet."""
    if "current" not in _template_store:
        raise HTTPException(status_code=400, detail="No template uploaded yet")

    import openpyxl

    # Check if there are documents to extract from
    idx = _get_index()
    stats = idx.describe_index_stats()
    ns = stats.get("namespaces", {})
    knowledge_count = ns.get(NAMESPACE_KNOWLEDGE, {}).get("vector_count", 0)

    # Fetch live property context (if any)
    property_context = _get_active_property_context()
    property_brief_chunk: str | None = None
    if property_context and property_context.property_brief:
        property_brief_chunk = f"[Source: NYC Live Property Data]\n{property_context.property_brief}"

    if knowledge_count == 0 and not property_brief_chunk:
        return {"updates": {}, "message": "No documents uploaded. Upload source documents first."}

    wb = openpyxl.load_workbook(io.BytesIO(_template_store["current"]["bytes"]), data_only=True)
    wb_f = openpyxl.load_workbook(io.BytesIO(_template_store["current"]["bytes"]), data_only=False)

    all_updates: dict[str, dict] = {}
    all_sources: dict[str, dict[str, str]] = {}
    all_confidence: dict[str, dict[str, str]] = {}

    ROWS_PER_BATCH = 40

    for name in wb.sheetnames:
        ws = wb[name]
        ws_f_s = wb_f[name]
        max_r = ws.max_row or 0
        max_c = ws.max_column or 0

        if max_r == 0 or max_c == 0:
            logging.info(f"  ⏭ Skipping sheet '{name}' (empty)")
            continue

        labels: list[str] = []

        # Build a grid: grid[r][c] → (display_value, is_formula, is_empty)
        grid: dict[int, dict[int, tuple]] = {}
        has_editable_empty = False
        for r in range(1, max_r + 1):
            grid[r] = {}
            for c in range(1, max_c + 1):
                val = ws.cell(r, c).value
                fml = ws_f_s.cell(r, c).value
                is_formula = isinstance(fml, str) and fml.startswith("=")
                if is_formula:
                    grid[r][c] = (None, True, False)
                elif val is None:
                    grid[r][c] = (None, False, True)
                    has_editable_empty = True
                else:
                    safe = _safe_cell_value(val)
                    grid[r][c] = (safe, False, False)
                    if isinstance(val, str) and not val.replace(".", "").replace("-", "").replace(",", "").replace(" ", "").isdigit():
                        labels.append(val)

        if not has_editable_empty:
            logging.info(f"  ⏭ Skipping sheet '{name}' (no empty editable cells)")
            continue

        # Detect header row — scan rows 1-3 for the one with the most text cells
        header_row: dict[int, str] = {}
        header_row_num = 0
        best_text_count = 0
        for candidate_r in range(1, min(max_r + 1, 4)):
            text_count = 0
            candidate_headers: dict[int, str] = {}
            for c in range(1, max_c + 1):
                val, is_f, is_e = grid[candidate_r].get(c, (None, False, True))
                if isinstance(val, str) and val.strip():
                    text_count += 1
                    candidate_headers[c] = val.strip()
            if text_count > best_text_count:
                best_text_count = text_count
                header_row = candidate_headers
                header_row_num = candidate_r

        # Build row-by-row descriptions
        row_descriptions: list[str] = []
        header_line = ""
        if header_row:
            hdr_parts = []
            for c in range(1, max_c + 1):
                col_letter = _col_letter(c)
                hdr_val = header_row.get(c, "")
                hdr_parts.append(f"{col_letter}=\"{hdr_val}\"" if hdr_val else col_letter)
            header_line = f"Row {header_row_num} (header): {' | '.join(hdr_parts)}"
            row_descriptions.append(header_line)

        start_row = (header_row_num + 1) if header_row else 1
        for r in range(start_row, max_r + 1):
            parts: list[str] = []
            row_has_content = False
            row_label = None
            for c in range(1, min(max_c + 1, 4)):
                val, is_f, is_e = grid[r].get(c, (None, False, True))
                if isinstance(val, str) and val.strip():
                    row_label = val.strip()
                    break

            for c in range(1, max_c + 1):
                val, is_formula, is_empty = grid[r].get(c, (None, False, True))
                col_letter = _col_letter(c)
                coord = f"{col_letter}{r}"
                col_header = header_row.get(c, "")

                if is_formula:
                    parts.append(f"{coord}=[formula]")
                    row_has_content = True
                elif is_empty:
                    context_hint = f" ({col_header})" if col_header else ""
                    parts.append(f"{coord}=<empty>{context_hint}")
                else:
                    # Mark existing values with [current] so LLM knows to overwrite
                    if isinstance(val, (int, float)):
                        parts.append(f"{coord}={val} [current]")
                    else:
                        parts.append(f"{coord}={val}")
                    row_has_content = True

            if row_has_content or any(
                not grid[r].get(c, (None, False, True))[2]
                for c in range(1, max_c + 1)
            ):
                label_hint = f" [{row_label}]" if row_label else ""
                row_descriptions.append(f"Row {r}{label_hint}: {' | '.join(parts)}")

        if not row_descriptions:
            logging.info(f"  ⏭ Skipping sheet '{name}' (no content rows)")
            continue

        logging.info(f"  📊 Sheet '{name}': {len(row_descriptions)} rows, {len(labels)} labels")

        # ── RAG retrieval ──────────────────────────────────────────────
        LABELS_PER_QUERY = 8
        label_groups = [labels[i:i + LABELS_PER_QUERY] for i in range(0, max(len(labels), 1), LABELS_PER_QUERY)]
        all_matches: list[dict] = []
        seen_ids: set[str] = set()

        # Batch-embed all label-group queries in one API call
        query_texts = [
            (f"UAP underwriting {name}: " + " ".join(group)) if group else f"UAP underwriting {name}"
            for group in label_groups
        ]
        query_embeddings = get_embeddings_batch(query_texts, client=openai_client, embedding_model=EMBEDDING_MODEL)

        for query_embedding in query_embeddings:
            results = idx.query(
                vector=query_embedding,
                top_k=50,
                namespace=NAMESPACE_KNOWLEDGE,
                include_metadata=True,
            )
            for m in results.get("matches", []):
                mid = m.get("id", "")
                if mid not in seen_ids:
                    seen_ids.add(mid)
                    all_matches.append(m)

        all_matches.sort(key=lambda m: m.get("score", 0), reverse=True)

        source_chunks, source_name_lookup = _select_diversified_source_chunks(all_matches)

        # Prepend live property data as an additional source
        if property_brief_chunk:
            source_chunks.insert(0, property_brief_chunk)
            source_name_lookup["nyc live property data"] = "NYC Live Property Data"

        if not source_chunks:
            logging.info(f"  ⏭ Skipping sheet '{name}' (no source chunks)")
            continue

        logging.info(f"  📄 Sheet '{name}': {len(all_matches)} RAG matches → {len(source_chunks)} chunks from {len(source_name_lookup)} docs")

        context_text = "\n---\n".join(source_chunks)
        glossary_context = _pick_relevant_glossary_sections(labels, name)

        # ── Split rows into batches ────────────────────────────────────
        data_rows = [rd for rd in row_descriptions if not rd.startswith(f"Row {header_row_num} (header)")]
        batches = [data_rows[i:i + ROWS_PER_BATCH] for i in range(0, max(len(data_rows), 1), ROWS_PER_BATCH)]

        sheet_updates: dict[str, object] = {}
        sheet_sources: dict[str, str] = {}
        sheet_confidence: dict[str, str] = {}

        logging.info(f"  🔄 Sheet '{name}': {len(batches)} batch(es) of ≤{ROWS_PER_BATCH} rows")

        for batch_idx, batch_rows in enumerate(batches):
            batch_desc = (header_line + "\n" if header_line else "") + "\n".join(batch_rows)

            prompt = (
                f'You are filling a UAP underwriting Excel spreadsheet from source documents.\n\n'
                f'Sheet: "{name}" (batch {batch_idx + 1}/{len(batches)})\n\n'
                f'LAYOUT KEY:\n'
                f'  CellRef=Value — cell with its current value\n'
                f'  CellRef=<empty> — cell with no value, needs to be filled\n'
                f'  CellRef=Value [current] — cell with an existing value that SHOULD BE OVERWRITTEN '
                f'if the source documents contain data for it\n'
                f'  CellRef=[formula] — auto-calculated, NEVER fill these\n'
                f'  Row labels in [brackets] describe each row\'s purpose\n'
                f'  The header row defines what each column means\n\n'
                f'Spreadsheet layout:\n{batch_desc}\n\n'
                f'Source document excerpts:\n{context_text}\n\n'
                f'YOUR TASK:\n'
                f'1. FILL EVERY POSSIBLE CELL.  Check every <empty> cell AND every [current] cell.\n'
                f'   For [current] cells, OVERWRITE with the value from the source documents —\n'
                f'   the source documents are the authority, not the existing spreadsheet values.\n'
                f'2. Return a JSON object mapping cell refs (e.g. "B5") to:\n'
                f'   {{"value": <number_or_string>, "source": "<exact filename>", "confidence": "high|medium|low"}}\n'
                f'3. Use the exact source name from the [Source: ...] header of each excerpt.\n'
                f'4. Numbers should be plain (no $ or commas).  Text fields should be strings.\n'
                f'5. NEVER fill [formula] cells.\n'
                f'6. Confidence: "high" = explicitly stated in source, "medium" = reasonable inference, '
                f'   "low" = educated guess or industry standard assumption.\n'
                f'7. BE AGGRESSIVE — fill as many cells as possible.  It is better to fill a cell with '
                f'   medium/low confidence than to leave it empty.\n'
                f'8. NEVER fill a numeric cell with 0 as a placeholder or guess.  Only use 0 if the '
                f'   source document explicitly states the value is zero.'
            )

            sys_msg = (
                "You are an expert NYC real estate underwriting analyst.  You read source documents "
                "(rent rolls, T-12 operating statements, appraisals, offering memorandums, tax bills, "
                "surveys, environmental reports, financial projections) and live NYC property data "
                "(PLUTO, DOF valuations, ACRIS transactions, HPD violations, DOB jobs, ECB violations, "
                "DOF comparable sales, HPD litigations, FDNY vacate orders) and populate underwriting "
                "spreadsheet cells.  The SOURCE DOCUMENTS are the single source of truth — if a cell "
                "already has a value marked [current], you MUST overwrite it with the value from the "
                "source documents.  Live property data (source 'NYC Live Property Data') is authoritative "
                "for property info, zoning, valuations, taxes, violations, and transaction history.  "
                "Use column headers and row labels to determine which value belongs "
                "in which cell.  Return only valid JSON.\n\n" + glossary_context
            )

            try:
                response = anthropic_client.messages.create(
                    model=GENERATION_MODEL,
                    max_tokens=16384,
                    temperature=0.15,
                    system=sys_msg,
                    messages=[
                        {"role": "user", "content": prompt},
                    ],
                )
                result_text = response.content[0].text.strip()
                # Strip markdown code fences if present
                if result_text.startswith("```"):
                    first_nl = result_text.index("\n") if "\n" in result_text else len(result_text)
                    result_text = result_text[first_nl + 1:]
                    if result_text.rstrip().endswith("```"):
                        result_text = result_text.rstrip()[:-3].rstrip()
                # Extract just the JSON object
                brace_start = result_text.find("{")
                if brace_start >= 0:
                    depth = 0
                    for i, c in enumerate(result_text[brace_start:], brace_start):
                        if c == "{":
                            depth += 1
                        elif c == "}":
                            depth -= 1
                            if depth == 0:
                                result_text = result_text[brace_start:i + 1]
                                break
                updates = json.loads(result_text)
                if isinstance(updates, dict):
                    batch_count = 0
                    for raw_ref, payload in updates.items():
                        if not isinstance(raw_ref, str):
                            continue
                        ref = raw_ref.strip().upper()
                        value = payload
                        source_name = None
                        confidence = None
                        if isinstance(payload, dict):
                            value = payload.get("value")
                            raw_source_name = payload.get("source")
                            if isinstance(raw_source_name, str):
                                source_name = source_name_lookup.get(raw_source_name.strip().lower())
                            raw_confidence = payload.get("confidence")
                            if isinstance(raw_confidence, str) and raw_confidence.lower() in ("high", "medium", "low"):
                                confidence = raw_confidence.lower()

                        if not isinstance(value, (str, int, float, bool)) and value is not None:
                            continue

                        # Discard zero-value low-confidence fills — likely hallucinated placeholders
                        if value == 0 and (confidence or "medium") == "low":
                            continue

                        sheet_updates[ref] = value
                        sheet_confidence[ref] = confidence or "medium"
                        if source_name:
                            sheet_sources[ref] = source_name
                        elif len(source_name_lookup) == 1:
                            sheet_sources[ref] = next(iter(source_name_lookup.values()))
                        batch_count += 1

                    logging.info(f"    ✅ Batch {batch_idx + 1}/{len(batches)}: {batch_count} cells")
            except Exception as e:
                logging.warning(f"    ❌ Batch {batch_idx + 1}/{len(batches)} failed for '{name}': {e}")

        if sheet_updates:
            all_updates[name] = sheet_updates
        if sheet_sources:
            all_sources[name] = sheet_sources
        if sheet_confidence:
            all_confidence[name] = sheet_confidence
        logging.info(f"  📊 Sheet '{name}' total: {len(sheet_updates)} cells extracted")

    total_cells = sum(len(v) for v in all_updates.values())
    logging.info(f"🏁 RAG extraction complete: {total_cells} cells across {len(all_updates)} sheets")
    return {"updates": all_updates, "sources": all_sources, "confidence": all_confidence}


@app.post("/api/underwriting/recalculate")
async def recalculate_underwriting_formulas(req: UnderwritingUpdatesRequest):
    """Recalculate workbook formulas using current cell updates without mutating the workbook."""
    if "current" not in _template_store:
        raise HTTPException(status_code=400, detail="No template uploaded")

    template = _template_store["current"]
    formula_values, warnings = calculate_underwriting_formula_values(
        template.get("calc_context"),
        req.updates,
    )
    return {
        "formulaValues": formula_values,
        "warnings": [warning.to_dict() for warning in warnings],
    }


@app.post("/api/underwriting/download")
async def download_filled_template(req: UnderwritingUpdatesRequest):
    """Apply cell updates to the stored template and return the filled .xlsx file."""
    if "current" not in _template_store:
        raise HTTPException(status_code=400, detail="No template uploaded")

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(_template_store["current"]["bytes"]))

    updates = req.updates
    for sheet_name, cells in updates.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for ref, value in cells.items():
            try:
                ws[ref] = value
            except Exception:
                continue

    enable_workbook_recalculation(wb)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    base = os.path.splitext(_template_store["current"]["filename"])[0]
    filename = f"{base}_filled.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── TC201 – Income & Expense Schedule for Rent-Producing Properties ─────


@app.get("/api/tc201")
async def get_tc201():
    """Return the currently loaded TC201 data, or a blank template."""
    if "current" not in _tc201_store:
        return TC201Data().model_dump()
    return _tc201_store["current"].model_dump()


@app.put("/api/tc201")
async def update_tc201(data: dict):
    """Update individual fields of the current TC201 data."""
    if "current" not in _tc201_store:
        _tc201_store["current"] = TC201Data()

    current = _tc201_store["current"]
    current_dict = current.model_dump()

    for key, value in data.items():
        if key in current_dict:
            if key == "residential_occupancy" and isinstance(value, list):
                current_dict[key] = [
                    ResidentialOccupancy(**r) if isinstance(r, dict) else r for r in value
                ]
            elif key == "nonresidential_floors" and isinstance(value, list):
                current_dict[key] = [
                    NonresidentialFloor(**f) if isinstance(f, dict) else f for f in value
                ]
            elif key == "misc_expenses" and isinstance(value, list):
                current_dict[key] = [
                    MiscExpenseItem(**m) if isinstance(m, dict) else m for m in value
                ]
            else:
                current_dict[key] = value

    _tc201_store["current"] = TC201Data(**current_dict)
    logging.info(f"Updated TC201 data for BBL={_tc201_store['current'].bbl}")
    return _tc201_store["current"].model_dump()


@app.delete("/api/tc201")
async def clear_tc201():
    """Clear the current TC201 data."""
    _tc201_store.pop("current", None)
    return {"cleared": True}


@app.post("/api/tc201/fill-from-property")
async def fill_tc201_from_property():
    """Auto-fill TC201 from active property context + uploaded docs via RAG."""
    property_context = _get_active_property_context()

    primary = None
    raw: dict = {}

    if property_context:
        primary = property_context.lots_detail[0] if property_context.lots_detail else None
        raw = primary.raw if primary else {}

    # ── Step 1: Build TC201 from live property data ─────────────────
    borough_names = {"1": "Manhattan", "2": "Bronx", "3": "Brooklyn", "4": "Queens", "5": "Staten Island"}
    if property_context:
        borough_code = property_context.borough or (property_context.primary_bbl[0] if property_context.primary_bbl else "")
        borough_name = borough_names.get(borough_code, borough_code)

        res_occ: list[ResidentialOccupancy] = []
        units_res = raw.get("unitsres") or raw.get("residential_units") or property_context.units_total
        if units_res:
            res_occ.append(ResidentialOccupancy(
                occupancy_type="RENTED, REGULATED",
                number_of_units=int(units_res) if units_res else None,
            ))

        tc201 = TC201Data(
            assessment_year="2026/27",
            borough=borough_name,
            block=property_context.block,
            lot=property_context.lots[0] if property_context.lots else "",
            bbl=property_context.primary_bbl,
            is_condo="Y" if raw.get("bldg_class", "").startswith("R") else "N",
            reporting_period_from="01/01/2025",
            reporting_period_to="12/31/2025",
            residential_occupancy=res_occ,
        )
    else:
        tc201 = TC201Data()

    # ── Step 2: RAG extraction from uploaded documents ──────────────
    try:
        idx = _get_index()
        stats = idx.describe_index_stats()
        ns = stats.get("namespaces", {})
        knowledge_count = ns.get(NAMESPACE_KNOWLEDGE, {}).get("vector_count", 0)
        logging.info(f"TC201 RAG: index={active_index_name}, knowledge_count={knowledge_count}")

        if knowledge_count > 0:
            # Retrieve ALL chunks — just fetch everything in KnowledgeStore
            # Use multiple targeted queries to maximize recall
            queries = [
                "total income revenue rent association fee parking",
                "total expense operating insurance management utilities",
                "net operating income net income profit loss",
                "repairs maintenance elevator water sewer electricity",
                "fuel cleaning wages payroll advertising insurance",
            ]

            seen_ids: set[str] = set()
            all_chunks: list[str] = []

            for q in queries:
                q_emb = get_embedding(q, client=openai_client, embedding_model=EMBEDDING_MODEL)
                results = idx.query(
                    vector=q_emb, top_k=15,
                    namespace=NAMESPACE_KNOWLEDGE, include_metadata=True,
                )
                for m in results.get("matches", []):
                    vid = m.get("id", "")
                    if vid in seen_ids:
                        continue
                    seen_ids.add(vid)
                    chunk_text = m.get("metadata", {}).get("text", "")
                    if chunk_text.strip():
                        all_chunks.append(chunk_text)

            logging.info(f"TC201 RAG: retrieved {len(all_chunks)} unique chunks")

            # Also add property brief if available
            if property_context and property_context.property_brief:
                all_chunks.insert(0, f"[NYC Live Property Data]\n{property_context.property_brief}")

            if all_chunks:
                # Concatenate all chunks into one context block
                full_context = "\n\n---\n\n".join(all_chunks)

                prompt = f"""You are filling out NYC Form TC201 (Income & Expense Schedule for Rent-Producing Properties).

Below are the financial documents for this property. They may contain:
- A P&L statement (monthly or year-to-date)
- A consolidated income statement (quarterly or annual)
- Property data from NYC

DOCUMENTS:
{full_context}

TASK: Extract the financial data and map it to TC201 fields. 

IMPORTANT RULES:
1. Use the ANNUAL / FULL-YEAR totals, not monthly or partial-year figures.
2. If there are two reports (e.g., a partial-year P&L and a full-year consolidated statement), prefer the FULL-YEAR consolidated numbers.
3. The TC201 has PRIOR YEAR and CURRENT YEAR columns. If the documents only cover one year, put those numbers in the CURRENT year fields (suffix _current). Leave prior year fields out.
4. Numbers must be plain integers (no $ signs, no commas, no decimals). Round to nearest dollar.
5. ONLY include fields where you found actual data. Do NOT guess or fabricate numbers.
6. TC201 is an OPERATING income/expense form. EXCLUDE all financing, capital, and non-operating items (see EXCLUSION list below).

MAP the data to these exact field names:

INCOME (Part 6):
- income_residential_regulated_current / _prior: Regulated residential rent (stabilized, rent-controlled)
- income_residential_unregulated_current / _prior: Unregulated / market-rate residential rent
- income_residential_subtotal_current / _prior: Sum of regulated + unregulated residential
- income_office_current / _prior: Office rental income
- income_retail_current / _prior: Retail / storefront income
- income_parking_current / _prior: Garage / parking income
- income_storage_current / _prior: Storage income
- income_other_current / _prior: Other income not categorized above (cash payments, late fees, etc.)
- income_other_description: Text description of "other" income
- income_subtotal_current / _prior: Subtotal of all rental categories above
- income_operating_escalation_current / _prior: Operating cost escalation pass-throughs (CAM reimbursements)
- income_re_tax_escalation_current / _prior: Real estate tax escalation / tenant RE tax reimbursements
- income_utility_services_current / _prior: Utility reimbursements from tenants
- income_govt_subsidies_current / _prior: Government rent subsidies
- income_total_gross_current / _prior: TOTAL gross income (sum of everything)

EXPENSES (Part 7):
- expense_fuel_current / _prior: Heating fuel / gas
- expense_light_power_current / _prior: Electricity / light and power
- expense_cleaning_current / _prior: Cleaning contracts / janitorial
- expense_wages_current / _prior: Wages, payroll, super salary
- expense_repairs_current / _prior: Repairs and maintenance (elevator, building, etc.)
- expense_management_current / _prior: Management and administration fees
- expense_insurance_current / _prior: Property insurance (liability + property)
- expense_water_sewer_current / _prior: Water and sewer
- expense_advertising_current / _prior: Advertising
- expense_painting_current / _prior: Interior painting and decorating
- expense_leasing_ti_current / _prior: Amortized leasing / tenant improvement costs
- expense_misc_current / _prior: Miscellaneous expenses (total from Part 9)
- expense_before_taxes_current / _prior: Total expenses BEFORE real estate taxes
- expense_real_estate_taxes_current / _prior: Real estate taxes (before abatements)
- expense_total_current / _prior: TOTAL expenses (before taxes + real estate taxes)

NET (Part 8):
- net_before_re_taxes_current / _prior: Net income before real estate taxes (total income - expenses before taxes)
- net_after_re_taxes_current / _prior: Net income after real estate taxes (total income - total expenses)

MISC EXPENSES itemization (Part 9) — break out what's in "miscellaneous":
- misc_expenses: array of {{"item": "description", "amount": number}}
  Examples: legal fees, accounting fees, extermination, fire alarm, snow removal, supplies, internet, bank fees

STRICT CATEGORIZATION RULES — follow these EXACTLY:

STEP 1 — EXCLUDE NON-OPERATING ITEMS (these NEVER go on TC201):
The following are FINANCING, CAPITAL, or NON-OPERATING costs. They must be COMPLETELY EXCLUDED
from ALL TC201 expense fields. Do NOT put them anywhere — just skip them:
  - Mortgage Expense / Mortgage Payment / Mortgage Principal → EXCLUDE
  - Loan Interest / Interest Expense / Loan Payment → EXCLUDE
  - Depreciation / Amortization (of assets, not leasing TI) → EXCLUDE
  - Capital Expenditures / Capital Improvements → EXCLUDE
  - Debt Service / Debt Payment → EXCLUDE
  - Owner Draws / Distributions → EXCLUDE
  - meals / entertainment / travel / auto / tolls → EXCLUDE (personal/business, not building operating)
  - Income Tax / Federal Tax / State Tax → EXCLUDE (only PROPERTY/RE tax goes on TC201)

STEP 2 — CATEGORIZE EACH REMAINING LINE ITEM EXACTLY ONCE:
Every OPERATING line item must appear in EXACTLY ONE TC201 category.
NEVER double-count: if a P&L section header (e.g. "City Regulations and Requirements: $5,153"
or "Repairs and Maintenance: $10,794") contains sub-items, use ONLY the leaf-level sub-items.
IGNORE all section header totals — they are just subtotals of their children.
Also IGNORE "Total for Expenses" — it may include excluded non-operating items.

STEP 3 — EXPENSE CATEGORY DEFINITIONS:

expense_repairs — ONLY physical repair or maintenance work on the building:
  - "Elevator Maintenance and Repair", "Elevator Repairs" → expense_repairs
  - "Elevator" (service/maintenance contract) → expense_repairs
  - "Door & Lock Repairs", "Garage Repair" → expense_repairs
  - "Sprinkler", "Fire Sprinkler" (maintenance/repair) → expense_repairs
  - "Building Materials", "Building Repairs" → expense_repairs
  - "Repairs & Maintenance" (the P&L total for this section) → expense_repairs
  - BUT: if "Repairs & Maintenance" has sub-items, sum the sub-items, do NOT also add the header total

expense_misc — EVERYTHING OPERATING that does NOT fit a named TC201 category:
  - "Backflow Test", "Elevator Test/Inspection", "HPD Registration", "Local Law" inspections → misc
  - "Extermination" → misc
  - "Fire Alarm", "alarm" (monitoring, NOT physical repair) → misc
  - "Snow Removal" → misc
  - "Supplies", "Office Supplies" → misc
  - "Postage and Delivery" → misc
  - "Phone", "Telephone", "Internet" → misc
  - "Legal Fees", "Accounting Fees", "Professional Fees" → misc
  - "Bank Fees", "Bank Service Charges" → misc
  - "Other Expenses", "Miscellaneous Expense" → misc

expense_misc_current MUST EQUAL the sum of all misc_expenses item amounts.

Other mappings:
  - "Rental Income" (residential property) → income_residential_regulated or income_residential_unregulated
  - "Rental Income" (commercial property) → income_retail or income_office as appropriate
  - "Association Fee Income" or "Monthly Maintenance Fee" → income_residential_regulated
  - "Parking Income" or "Parking Spot" → income_parking
  - "Special Assessment Income" or "Contribution To Reserve Account" → income_other
  - "CAM Reimbursements", "CAM - Tenant Reimbursements" → income_operating_escalation
  - "RE Tax Reimbursements", "RE Taxes-Tenant Reimbursements" → income_re_tax_escalation
  - "Utility Reimbursements", "Utilities-Tenant reimbursements" → income_utility_services
  - "Super Salary", "Wages" → expense_wages
  - "Property Insurance" + "Liability" → expense_insurance (sum both)
  - "Electricity", "Utilities" (if single electric line) → expense_light_power
  - "Water & Sewer", "Water" → expense_water_sewer
  - "Management Fees", "Management Fee" → expense_management
  - "Cleaning" → expense_cleaning
  - "Advertising" → expense_advertising
  - "Property Taxes", "Property Tax", "RE Tax" → expense_real_estate_taxes

FINAL VERIFICATION — YOU MUST DO THIS BEFORE RETURNING:
1. Confirm you excluded ALL financing items (mortgage, interest, depreciation, debt service)
2. Confirm no section header totals were added alongside their sub-items (no double-counting)
3. Sum your individual expense fields: fuel + light + cleaning + wages + repairs + management + insurance + water + advertising + painting + leasing_ti + misc
4. Set expense_before_taxes_current to that exact sum
5. The net = income_total_gross - expense_before_taxes

Return ONLY a JSON object with the fields you found data for. No nulls, no empty strings."""

                try:
                    response = anthropic_client.messages.create(
                        model=GENERATION_MODEL,
                        max_tokens=16384,
                        temperature=0.0,
                        system="You are an expert NYC real estate accountant filling Form TC201. Extract and map financial data precisely. Return only a JSON object.",
                        messages=[
                            {"role": "user", "content": prompt},
                        ],
                    )
                    result_text = response.content[0].text.strip()
                    # Strip markdown code fences if present
                    if result_text.startswith("```"):
                        first_nl = result_text.index("\n") if "\n" in result_text else len(result_text)
                        result_text = result_text[first_nl + 1:]
                        if result_text.rstrip().endswith("```"):
                            result_text = result_text.rstrip()[:-3].rstrip()
                    # Extract just the JSON object (Claude may add text after it)
                    brace_start = result_text.find("{")
                    if brace_start >= 0:
                        depth = 0
                        for i, c in enumerate(result_text[brace_start:], brace_start):
                            if c == "{":
                                depth += 1
                            elif c == "}":
                                depth -= 1
                                if depth == 0:
                                    result_text = result_text[brace_start:i + 1]
                                    break
                    logging.info(f"TC201 RAG response (first 1000 chars): {result_text[:1000]}")
                    rag_values = json.loads(result_text)
                    logging.info(f"TC201 RAG parsed {len(rag_values)} keys: {list(rag_values.keys())}")

                    if isinstance(rag_values, dict):
                        tc201_dict = tc201.model_dump()

                        # Handle nested list fields
                        for list_key, model_cls in [
                            ("residential_occupancy", ResidentialOccupancy),
                            ("nonresidential_floors", NonresidentialFloor),
                            ("misc_expenses", MiscExpenseItem),
                        ]:
                            if list_key in rag_values and isinstance(rag_values[list_key], list):
                                items = []
                                for item in rag_values[list_key]:
                                    if isinstance(item, dict):
                                        items.append(model_cls(**item))
                                if items:
                                    tc201_dict[list_key] = items

                        # Merge scalar fields — RAG values override empty/null fields
                        skip = {"residential_occupancy", "nonresidential_floors", "misc_expenses", "filename", "assessment_year"}
                        for key, value in rag_values.items():
                            if key in skip or key not in tc201_dict or value is None:
                                continue
                            current_val = tc201_dict.get(key)
                            if current_val is None or current_val == "" or current_val == []:
                                tc201_dict[key] = value

                        tc201 = TC201Data(**tc201_dict)

                        # ── Post-processing: enforce math consistency ──
                        # Recompute misc total from itemized list
                        if tc201.misc_expenses:
                            misc_sum = sum(m.amount for m in tc201.misc_expenses if m.amount)
                            tc201.expense_misc_current = misc_sum

                        # Recompute income subtotals
                        income_components = [
                            tc201.income_residential_regulated_current,
                            tc201.income_residential_unregulated_current,
                            tc201.income_office_current,
                            tc201.income_retail_current,
                            tc201.income_loft_current,
                            tc201.income_factory_current,
                            tc201.income_warehouse_current,
                            tc201.income_storage_current,
                            tc201.income_parking_current,
                        ]
                        computed_subtotal = sum(v for v in income_components if v)
                        if computed_subtotal > 0:
                            tc201.income_subtotal_current = computed_subtotal

                        res_components = [
                            tc201.income_residential_regulated_current,
                            tc201.income_residential_unregulated_current,
                        ]
                        computed_res = sum(v for v in res_components if v)
                        if computed_res > 0:
                            tc201.income_residential_subtotal_current = computed_res

                        # Recompute gross income
                        gross_addons = [
                            tc201.income_subtotal_current,
                            tc201.income_owner_occupied_current,
                            tc201.income_operating_escalation_current,
                            tc201.income_re_tax_escalation_current,
                            tc201.income_utility_services_current,
                            tc201.income_other_services_current,
                            tc201.income_govt_subsidies_current,
                            tc201.income_signage_current,
                            tc201.income_cell_towers_current,
                            tc201.income_other_current,
                        ]
                        computed_gross = sum(v for v in gross_addons if v)
                        if computed_gross > 0:
                            tc201.income_total_gross_current = computed_gross

                        # Recompute expense totals
                        expense_components = [
                            tc201.expense_fuel_current,
                            tc201.expense_light_power_current,
                            tc201.expense_cleaning_current,
                            tc201.expense_wages_current,
                            tc201.expense_repairs_current,
                            tc201.expense_management_current,
                            tc201.expense_insurance_current,
                            tc201.expense_water_sewer_current,
                            tc201.expense_advertising_current,
                            tc201.expense_painting_current,
                            tc201.expense_leasing_ti_current,
                            tc201.expense_misc_current,
                        ]
                        computed_expense = sum(v for v in expense_components if v)
                        if computed_expense > 0:
                            tc201.expense_before_taxes_current = computed_expense

                        if tc201.expense_real_estate_taxes_current:
                            tc201.expense_total_current = computed_expense + (tc201.expense_real_estate_taxes_current or 0)
                        else:
                            tc201.expense_total_current = computed_expense

                        # Recompute net
                        if tc201.income_total_gross_current and tc201.expense_before_taxes_current:
                            tc201.net_before_re_taxes_current = tc201.income_total_gross_current - tc201.expense_before_taxes_current
                        if tc201.income_total_gross_current and tc201.expense_total_current:
                            tc201.net_after_re_taxes_current = tc201.income_total_gross_current - tc201.expense_total_current

                        logging.info(f"TC201 post-processing: income={tc201.income_total_gross_current}, expense={tc201.expense_before_taxes_current}, net={tc201.net_before_re_taxes_current}")
                        logging.info(f"TC201 RAG extraction merged {len(rag_values)} fields")

                except Exception as e:
                    logging.warning(f"TC201 RAG extraction failed (non-fatal): {e}")

    except Exception as e:
        logging.warning(f"TC201 document lookup failed (non-fatal): {e}")

    _tc201_store["current"] = tc201
    has_property = bool(property_context)
    logging.info(f"TC201 filled: BBL={tc201.bbl}, property={'yes' if has_property else 'no'}")
    return tc201.model_dump()


@app.get("/api/tc201/download")
async def download_tc201_pdf():
    """Generate and download the current TC201 data as a PDF."""
    if "current" not in _tc201_store:
        raise HTTPException(status_code=400, detail="No TC201 data to download")

    tc201 = _tc201_store["current"]
    pdf_bytes = _generate_tc201_pdf(tc201)

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="TC201_{tc201.bbl or "blank"}_{tc201.assessment_year.replace("/", "-")}.pdf"'},
    )


@app.get("/api/tc201/template")
async def get_tc201_template():
    """Serve the raw TC201 template PDF for client-side filling."""
    if not os.path.exists(TC201_TEMPLATE_PATH):
        raise HTTPException(status_code=404, detail="TC201 template not found")
    with open(TC201_TEMPLATE_PATH, "rb") as f:
        content = f.read()
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="tc201_template.pdf"'},
    )


@app.post("/api/tc201/upload-template")
async def upload_tc201_template(file: UploadFile = File(...)):
    """Replace the TC201 template PDF with an uploaded file."""
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="File must be a PDF")
    content = await file.read()
    if len(content) < 1000:
        raise HTTPException(status_code=400, detail="File too small to be a valid PDF")
    with open(TC201_TEMPLATE_PATH, "wb") as f:
        f.write(content)
    from pypdf import PdfReader as _PR
    pages = len(_PR(TC201_TEMPLATE_PATH).pages)
    return {"ok": True, "size": len(content), "pages": pages}


# ── TC201 PDF fill (overlay text on official form) ──────────────────────

TC201_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "tc201_template.pdf")


def _fmt_currency(val: float | None) -> str:
    if val is None:
        return ""
    return f"{val:,.0f}"


def _generate_tc201_pdf(data: TC201Data) -> bytes:
    """Fill the official TC201 AcroForm fields and return PDF bytes."""
    from pypdf import PdfReader, PdfWriter

    reader = PdfReader(TC201_TEMPLATE_PATH)
    writer = PdfWriter()
    writer.append(reader)

    def v(val) -> str:
        """Stringify a value for a form field; skip None."""
        if val is None:
            return ""
        return str(val)

    def cur(val) -> str:
        """Format number as currency string (no $) for form fields."""
        if val is None:
            return ""
        return f"{val:,.0f}"

    # Build field-value mapping from TC201Data → AcroForm field names
    fields: dict[str, str] = {}

    # ── Part 1: Property Identification ──
    fields["Borough"] = v(data.borough)
    fields["Block"] = v(data.block)
    fields["Lot"] = v(data.lot)
    fields["GROUP"] = v(data.tax_commission_group_no)
    fields["REP_TC_GROUP_NUMBER"] = v(data.tax_commission_group_no)
    fields["is_condo_cover_all_tc109"] = v(data.is_condo)
    fields["is_cover_more_one_lot"] = v(data.covers_multiple_lots)
    if data.total_lots is not None:
        fields["total_lot_number"] = v(data.total_lots)
    fields["is_schedule_for_intire_lots"] = v(data.covers_entire_lot)
    if data.related_lots:
        fields["lots_1"] = v(data.related_lots)
    if hasattr(data, 'describe_omission') and data.describe_omission:
        fields["describe_omission"] = v(data.describe_omission)

    # ── Part 2: Reporting period ──
    if data.reporting_period_from:
        parts = data.reporting_period_from.replace("-", "/").split("/")
        if len(parts) >= 3:
            fields["rp_from_month"] = parts[0]
            fields["rp_from_day"] = parts[1]
            fields["rp_from_year"] = parts[2]
    if data.reporting_period_to:
        parts = data.reporting_period_to.replace("-", "/").split("/")
        if len(parts) >= 3:
            fields["rp_to_month"] = parts[0]
            fields["rp_to_day"] = parts[1]
            fields["rp_to_year"] = parts[2]
    # Accounting basis checkboxes
    if data.accounting_basis.lower().startswith("cash"):
        fields["cb_Accbasis_cash"] = "/On"
        fields["cb_Accbasis_Accrual"] = "/Off"
    elif data.accounting_basis.lower().startswith("accrual"):
        fields["cb_Accbasis_Accrual"] = "/On"
        fields["cb_Accbasis_cash"] = "/Off"
    if data.accounting_basis_changed.upper() == "Y":
        fields["cb_Accbasis_changed_yes"] = "/On"
        fields["cb_Accbasis_changed_no"] = "/Off"
    elif data.accounting_basis_changed.upper() == "N":
        fields["cb_Accbasis_changed_no"] = "/On"
        fields["cb_Accbasis_changed_yes"] = "/Off"

    # ── Part 3: Residential Occupancy ──
    occ_map = {
        "RENTED, REGULATED": ("regulated_units_number", "regulated_monthly_rent"),
        "RENTED, UNREGULATED": ("unregulated_units_number", "unregulated_monthly_rent"),
        "OWNER OCCUPIED": ("owner_occupied_units_number", "owner_occupied_monthly_rent"),
        "VACANT": ("vacant_units_number", "vacant_monthly_rent"),
    }
    total_units = 0
    total_rent = 0.0
    for r in data.residential_occupancy:
        key = r.occupancy_type.upper().strip()
        for k, (units_field, rent_field) in occ_map.items():
            if key.startswith(k[:6]):
                if r.number_of_units is not None:
                    fields[units_field] = v(r.number_of_units)
                    total_units += r.number_of_units
                if r.monthly_rent is not None:
                    fields[rent_field] = cur(r.monthly_rent)
                    total_rent += r.monthly_rent
                break
    if total_units:
        fields["total_units_number"] = v(total_units)
    if total_rent:
        fields["total_monthly_rent"] = cur(total_rent)
    if data.rent_includes_recurring_charges:
        fields["is_include_all_charges"] = v(data.rent_includes_recurring_charges)

    # ── Part 4: Nonresidential Occupancy ──
    floor_prefix_map = {
        "FLOOR 3": "fl3", "3": "fl3",
        "2ND": "fl2", "2": "fl2",
        "1ST": "fl1", "1": "fl1",
        "BASEMENT": "base", "B": "base",
    }
    for nf in data.nonresidential_floors:
        key = nf.floor.upper().strip()
        prefix = None
        for k, p in floor_prefix_map.items():
            if key.startswith(k):
                prefix = p
                break
        if prefix is None:
            continue
        if nf.applicant_related_sqft is not None:
            fields[f"{prefix}_applicant"] = cur(nf.applicant_related_sqft)
        if nf.rented_sqft is not None:
            fields[f"{prefix}_rented"] = cur(nf.rented_sqft)
        if nf.vacant_sqft is not None:
            fields[f"{prefix}_vacant"] = cur(nf.vacant_sqft)
        if nf.gross_sqft is not None:
            fields[f"{prefix}_total"] = cur(nf.gross_sqft)

    # ── Part 5: Lease info ──
    fields["is_ground_lease_y"] = v(data.entire_lot_leased)
    if data.lease_type:
        lt = data.lease_type.lower()
        fields["cb_lease_type_gross"] = "/On" if "gross" in lt else "/Off"
        fields["cb_lease_type_net"] = "/On" if lt == "net" else "/Off"
        fields["cb_lease_type_ground"] = "/On" if "ground" in lt else "/Off"
    if data.applicant_receives_rental_income:
        fields["is_lessee_receives_rental_income"] = v(data.applicant_receives_rental_income)
    if data.lessor:
        fields["LESSOR"] = v(data.lessor)
    if data.lessee:
        fields["LESSEE"] = v(data.lessee)
    if data.lease_from:
        parts = data.lease_from.replace("-", "/").split("/")
        if len(parts) >= 2:
            fields["term_of_lease_from_month"] = parts[0]
            fields["term_of_lease_from_year"] = parts[-1]
    if data.lease_to:
        parts = data.lease_to.replace("-", "/").split("/")
        if len(parts) >= 2:
            fields["term_of_lease_to_month"] = parts[0]
            fields["term_of_lease_to_year"] = parts[-1]
    if data.annual_rent is not None:
        fields["annual_rent"] = cur(data.annual_rent)
    if data.additional_sums is not None:
        fields["additional_sums"] = cur(data.additional_sums)

    # ── Part 6: Income (current year) ──
    fields["Regulated"] = cur(data.income_residential_regulated_current)
    fields["Unregulated"] = cur(data.income_residential_unregulated_current)
    fields["SubtotalResidentialIncome"] = cur(data.income_residential_subtotal_current)
    fields["Office"] = cur(data.income_office_current)
    fields["Retail_Tenants"] = cur(data.income_retail_current)
    fields["Loft"] = cur(data.income_loft_current)
    fields["Factory"] = cur(data.income_factory_current)
    fields["Warehouse"] = cur(data.income_warehouse_current)
    fields["Storage"] = cur(data.income_storage_current)
    fields["Garages"] = cur(data.income_parking_current)
    fields["Subtotal"] = cur(data.income_subtotal_current)
    fields["Owner_Related"] = cur(data.income_owner_occupied_current)
    fields["Operating"] = cur(data.income_operating_escalation_current)
    fields["Tax_Escalation"] = cur(data.income_re_tax_escalation_current)
    fields["Utility_Services"] = cur(data.income_utility_services_current)
    fields["Other_Services"] = cur(data.income_other_services_current)
    fields["Rent_Sub"] = cur(data.income_govt_subsidies_current)
    fields["Signage"] = cur(data.income_signage_current)
    fields["Cell_Towers"] = cur(data.income_cell_towers_current)
    fields["Other"] = cur(data.income_other_current)
    if data.income_other_description:
        fields["INCOME_Other_NAME"] = v(data.income_other_description)
    fields["Total_inc_Est"] = cur(data.income_total_gross_current)

    # ── Part 6: Income (prior year — prefixed with 'p') ──
    fields["pRegulated"] = cur(data.income_residential_regulated_prior)
    fields["pUnregulated"] = cur(data.income_residential_unregulated_prior)
    fields["pSubtotalResidentialIncome"] = cur(data.income_residential_subtotal_prior)
    fields["pOffice"] = cur(data.income_office_prior)
    fields["pRetail_Tenants"] = cur(data.income_retail_prior)
    fields["pLoft"] = cur(data.income_loft_prior)
    fields["pFactory"] = cur(data.income_factory_prior)
    fields["pWarehouse"] = cur(data.income_warehouse_prior)
    fields["pStorage"] = cur(data.income_storage_prior)
    fields["pGarages"] = cur(data.income_parking_prior)
    fields["Psubtotal"] = cur(data.income_subtotal_prior)
    fields["pOwner_Related"] = cur(data.income_owner_occupied_prior)
    fields["pOperating"] = cur(data.income_operating_escalation_prior)
    fields["pTax_Escalation"] = cur(data.income_re_tax_escalation_prior)
    fields["pUtility_Services"] = cur(data.income_utility_services_prior)
    fields["pOther_Services"] = cur(data.income_other_services_prior)
    fields["pRent_Sub"] = cur(data.income_govt_subsidies_prior)
    fields["pSignage"] = cur(data.income_signage_prior)
    fields["pCell_Towers"] = cur(data.income_cell_towers_prior)
    fields["pOther"] = cur(data.income_other_prior)
    fields["pTotal_inc_Est"] = cur(data.income_total_gross_prior)

    # ── Part 7: Expenses (current year) ──
    fields["Fuel"] = cur(data.expense_fuel_current)
    fields["Light"] = cur(data.expense_light_power_current)
    fields["Cleaning"] = cur(data.expense_cleaning_current)
    fields["Wages"] = cur(data.expense_wages_current)
    fields["Repairs"] = cur(data.expense_repairs_current)
    fields["Management"] = cur(data.expense_management_current)
    fields["Insurance"] = cur(data.expense_insurance_current)
    fields["Water"] = cur(data.expense_water_sewer_current)
    fields["Advertising"] = cur(data.expense_advertising_current)
    fields["Interior"] = cur(data.expense_painting_current)
    fields["Amortized"] = cur(data.expense_leasing_ti_current)
    fields["Misc"] = cur(data.expense_misc_current)
    fields["SubTot_Expense"] = cur(data.expense_before_taxes_current)
    fields["Real_Taxes"] = cur(data.expense_real_estate_taxes_current)
    fields["Total_Expense"] = cur(data.expense_total_current)

    # ── Part 7: Expenses (prior year) ──
    fields["pFuel"] = cur(data.expense_fuel_prior)
    fields["pLight"] = cur(data.expense_light_power_prior)
    fields["pCleaning"] = cur(data.expense_cleaning_prior)
    fields["pWages"] = cur(data.expense_wages_prior)
    fields["pRepairs"] = cur(data.expense_repairs_prior)
    fields["pManagement"] = cur(data.expense_management_prior)
    fields["pInsurance"] = cur(data.expense_insurance_prior)
    fields["pWater"] = cur(data.expense_water_sewer_prior)
    fields["pAdvertising"] = cur(data.expense_advertising_prior)
    fields["pInterior"] = cur(data.expense_painting_prior)
    fields["pAmortized"] = cur(data.expense_leasing_ti_prior)
    fields["pMisc"] = cur(data.expense_misc_prior)
    fields["pSubTot_Expense"] = cur(data.expense_before_taxes_prior)
    fields["pReal_Taxes"] = cur(data.expense_real_estate_taxes_prior)
    fields["pTotal_Expense"] = cur(data.expense_total_prior)

    # ── Part 8: Net ──
    fields["before_taxes"] = cur(data.net_before_re_taxes_current)
    fields["After_taxes"] = cur(data.net_after_re_taxes_current)
    fields["pbefore_taxes"] = cur(data.net_before_re_taxes_prior)
    fields["pAfter_taxes"] = cur(data.net_after_re_taxes_prior)

    # ── Part 9: Misc expense itemization ──
    for i, m in enumerate(data.misc_expenses[:8]):
        fields[f"item{i + 1}"] = v(m.item) if m.item else ""
        fields[f"amount{i + 1}"] = cur(m.amount) if m.amount is not None else ""

    # ── Part 10: Tenants' electricity ──
    if data.tenants_electricity_from_applicant:
        fields["is_tenants_from_applicant"] = v(data.tenants_electricity_from_applicant)
    if data.tenants_electricity_separate_charge:
        fields["is_separate_charge"] = v(data.tenants_electricity_separate_charge)

    # Strip empty values — only fill fields that have data
    fields = {k: val for k, val in fields.items() if val}

    # First, clear ALL form fields to remove residual data from the template
    all_field_names = set()
    for page in reader.pages:
        if "/Annots" in page:
            for annot in page["/Annots"]:
                annot_obj = annot.get_object()
                if annot_obj.get("/T"):
                    all_field_names.add(str(annot_obj["/T"]))
    blank_fields = {name: "" for name in all_field_names if name not in fields}

    # Write blank fields first (clear residual data), then our filled fields
    for page_num in range(len(writer.pages)):
        writer.update_page_form_field_values(writer.pages[page_num], blank_fields)
        writer.update_page_form_field_values(writer.pages[page_num], fields)

    output = io.BytesIO()
    writer.write(output)
    return output.getvalue()


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        "main:app",
        host=os.getenv("BACKEND_HOST", "0.0.0.0"),
        port=int(os.getenv("BACKEND_PORT", "8000")),
        reload=True,
    )

